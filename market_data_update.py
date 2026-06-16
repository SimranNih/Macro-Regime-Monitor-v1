import yfinance as yf
from openpyxl import load_workbook

EXCEL_FILE = r"C:\Users\Simran Nihalani\OneDrive\Documents\Models\Macro Regime Monitor\Macro dashboard.xlsm"
RAW_SHEET = "Master Raw Data"

#Making the ticker map

ticker_map = {
    "US": {
        "equity": "^GSPC",
        "fx": None
    },

    "Eurozone": {
        "equity": "^STOXX50E",
        "fx": "EURUSD=X"
    },

    "UK": {
        "equity": "^FTSE",
        "fx": "GBPUSD=X"
    },

    "Japan": {
        "equity": "^N225",
        "fx": "JPY=X"
    },

    "Australia": {
        "equity": "^AXJO",
        "fx": "AUDUSD=X"
    },

    "Switzerland": {
        "equity": "^SSMI",
        "fx": "CHF=X"
    },

    "Norway": {
        "equity": "OBX.OL",
        "fx": "NOK=X"
    },

    "Canada": {
        "equity": "^GSPTSE",
        "fx": "CAD=X"
    },

    "China": {
        "equity": "000300.SS",
        "fx": "CNY=X"
    },

    "India": {
        "equity": "^NSEI",
        "fx": "INR=X"
    },

    "South Korea": {
        "equity": "^KS11",
        "fx": "KRW=X"
    },

    "Taiwan": {
        "equity": "^TWII",
        "fx": "TWD=X"
    },

    "Indonesia": {
        "equity": "^JKSE",
        "fx": "IDR=X"
    },

    "Brazil": {
        "equity": "^BVSP",
        "fx": "BRL=X"
    },

    "Mexico": {
        "equity": "^MXX",
        "fx": "MXN=X"

    },

    "Chile": {
        "equity": "^IPSA",
        "fx": "USDCLP=X"
    },

    "South Africa": {
        "equity": "^J203.JO",
        "fx": "ZAR=X"
    },

    "Turkey": {
        "equity": "XU100.IS",
        "fx": "TRY=X"
    },

    "Saudi Arabia": {
        "equity": "^TASI.SR",
        "fx": "SAR=X"
    }
}

#fx pairs that need inversion
invert_fx = {
    "EURUSD=X",
    "GBPUSD=X",
    "AUDUSD=X"
}

#opening excel

wb=load_workbook(EXCEL_FILE, keep_vba=True) 
ws=wb[RAW_SHEET]

#updating data

for row in range(2,ws.max_row +1):
    country=ws[f"A{row}"].value
    if country not in ticker_map:
        continue

    try:

        equity_ticker = ticker_map[country]["equity"]

        eq = yf.download(equity_ticker,
                         period="18mo",
                         auto_adjust=True,
                         progress=False
        )
        close=eq["Close"]

        equity_current = float(close.iloc[-1])
        equity_3m = float(close.iloc[-63])
        equity_12m = float(close.iloc[-252])

        ws[f"V{row}"] = round(equity_current,2)
        ws[f"W{row}"] = round(equity_3m,2)
        ws[f"X{row}"] = round(equity_12m,2)

        #FX Data

        fx_ticker = ticker_map[country]["fx"]

        if fx_ticker:

            fx = yf.download(
                fx_ticker,
                period="18mo",
                auto_adjust=True,
                progress=False
            )

            fx_close = fx["Close"]

            fx_current = float(fx_close.iloc[-1])
            fx_3m = float(fx_close.iloc[-63])
            fx_12m = float(fx_close.iloc[-252])

            if fx_ticker in invert_fx:

                fx_current = 1/fx_current
                fx_3m = 1/fx_3m
                fx_12m = 1/fx_12m

            ws[f"Y{row}"] = round(fx_current,4)
            ws[f"Z{row}"] = round(fx_3m,4)
            ws[f"AA{row}"] = round(fx_12m,4)

        print (f"Updated {country}")
    except Exception as e:
        print(f"Error updating {country}: {e}")

wb.save(EXCEL_FILE)

print("Market Data Update Complete")




